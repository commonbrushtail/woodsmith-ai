"""
Read the xlsx product data and generate a SQL migration to seed:
- variation_groups (ขนาด, สี, เกรด, ความหนา)
- variation_entries (unique values per group)
- products (~39 grouped products)
- product_variation_links (linking products to their variation entries)
"""

import openpyxl
import uuid
import re
import json

XLSX_PATH = 'Vanachai website map 040326/Vanachai website map/Website map - photo.xlsx'
OUTPUT_PATH = 'supabase/migrations/042_seed_products_from_xlsx.sql'

# Category -> product_type mapping
CATEGORY_TYPE = {
    'โต๊ะ': 'decoration',
    'ประตู': 'decoration',
    'ไม้โพลีเอสเตอร์': 'construction',
    'MDF': 'construction',
    'PB': 'construction',
    'OSB': 'construction',
    'ไม้อัด': 'construction',
    'ไม้บันได': 'decoration',
    'ไม้แบบ': 'construction',
    'ไม้พื้น': 'decoration',
    'ไม้ฝาตกแต่ง': 'decoration',
}

# Sheet columns: 0=name_th, 1=name_en, 2=size, 3=description, 4=grade, 5=other, 6=sku, 7=color, 8=thickness, 9=image

def slugify(text):
    """Convert Thai/English text to URL-friendly slug."""
    # Map Thai category names to English slugs
    slug_map = {
        'โต๊ะ': 'table',
        'ประตู': 'door',
        'ไม้โพลีเอสเตอร์': 'polyester-board',
        'MDF': 'mdf',
        'PB': 'pb',
        'OSB': 'osb',
        'ไม้อัด': 'plywood',
        'ไม้บันได': 'stair',
        'ไม้แบบ': 'shuttering-board',
        'ไม้พื้น': 'flooring',
        'ไม้ฝาตกแต่ง': 'decorative-panel',
        'ประตูเมลามีน': 'melamine-door',
        'ประตูเมลามีนพร้อมวงกบ': 'melamine-door-with-frame',
        'ประตูเมลามีนกันน้ำ': 'ultra-melamine-door',
        'ประตูลูกฟัก': 'hdf-door',
        'ประตูลูกฟักพร้อมวงกบ': 'hdf-door-with-frame',
        'ประตูHMR': 'hmr-door',
        'ประตูปิดPVC': 'pvc-door',
        'ประตูไม้อัด': 'plywood-door',
        'ประตูสักประดิษฐ์': 'recompose-teak-door',
        'หน้าประตู': 'doorskin',
        'ไม้ PB': 'pb-board',
        'ไม้ PB ปิดผิวเมลามีน': 'pb-melamine',
        'ไม้ OSB2 ภายใน': 'osb2-interior',
        'ไม้ OSB2 ทนชื้น': 'osb2-moisture',
        'ไม้ OSB2 ภายนอก': 'osb2-exterior',
        'ไม้ OSB3 ภายใน': 'osb3-interior',
        'ไม้ OSB3 ภายนอก': 'osb3-exterior',
        'ไม้พื้นลามิเนต': 'laminate-flooring',
        'ไม้พื้นลามิเนตปลอดสารฟอมัลดีไฮด์': 'formaldehyde-free-laminate',
        'ไม้พื้น Veneer Pattern': 'veneer-pattern-flooring',
        'ไม้พื้น Veneer': 'veneer-flooring',
        'ไม้พื้น Hybrid HMR': 'hybrid-hmr-flooring',
        'ไม้พื้น Hybrid Ultra': 'hybrid-ultra-flooring',
        'ไม้ MDF': 'mdf-board',
        'ไม้ MDF ปิดผิวเมลามีน': 'mdf-melamine',
        'ไม้ MDF HMR V313': 'mdf-hmr-v313',
        'ไม้ MDF HMR V70': 'mdf-hmr-v70',
        'ไม้ฝาตกแต่ง US': 'decorative-panel-us',
        'ไม้ฝาตกแต่ง AR': 'decorative-panel-ar',
        'ไม้ฝา HDF ปิดPVC': 'hdf-pvc-panel',
        'ไม้ฝา LVL ลายSaw Cut': 'lvl-saw-cut-panel',
        'ไม้ฝา LVL ลายสน': 'lvl-pine-panel',
        'ไม้อัด': 'plywood-board',
        'ไม้อัด ภายใน': 'plywood-interior',
        'ไม้อัด ภายนอก': 'plywood-exterior',
        'ไม้แบบ': 'shuttering',
        'ไม้บันได': 'stair-wood',
    }
    if text in slug_map:
        return slug_map[text]
    # Fallback: use English name or transliterate
    s = text.lower().strip()
    s = re.sub(r'[^a-z0-9\s-]', '', s)
    s = re.sub(r'[\s]+', '-', s)
    s = re.sub(r'-+', '-', s).strip('-')
    return s or f'product-{uuid.uuid4().hex[:8]}'


def escape_sql(val):
    """Escape single quotes for SQL."""
    if val is None:
        return 'NULL'
    return "'" + str(val).replace("'", "''") + "'"


def main():
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)

    product_sheets = ['โต๊ะ', 'ประตู', 'ไม้โพลีเอสเตอร์', 'MDF', 'PB', 'OSB',
                      'ไม้อัด', 'ไม้บันได', 'ไม้แบบ', 'ไม้พื้น', 'ไม้ฝาตกแต่ง']

    # Collect all products grouped by (sheet, product_name)
    all_products = {}  # key: (sheet, product_name) -> { name_en, rows: [...] }

    for sheet_name in product_sheets:
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            name_th = row[0]
            if name_th is None:
                continue
            name_th = str(name_th).strip()
            # Skip remark rows (start with *)
            if name_th.startswith('*'):
                continue
            key = (sheet_name, name_th)
            if key not in all_products:
                all_products[key] = {
                    'name_th': name_th,
                    'name_en': str(row[1]).strip() if row[1] else '',
                    'category': sheet_name,
                    'type': CATEGORY_TYPE.get(sheet_name, 'construction'),
                    'sizes': set(),
                    'colors': set(),
                    'grades': set(),
                    'thicknesses': set(),
                    'descriptions': set(),
                    'skus': set(),
                    'images': set(),
                }
            p = all_products[key]
            if row[2]: p['sizes'].add(str(row[2]).strip())
            if row[7]: p['colors'].add(str(row[7]).strip())
            if row[4]: p['grades'].add(str(row[4]).strip())
            if row[8]: p['thicknesses'].add(str(row[8]).strip())
            if row[3]: p['descriptions'].add(str(row[3]).strip())
            if row[6]: p['skus'].add(str(row[6]).strip())
            if row[9]: p['images'].add(str(row[9]).strip())

    # Create variation groups with fixed UUIDs
    vg_size_id = '20000000-0000-0000-0000-000000000001'
    vg_color_id = '20000000-0000-0000-0000-000000000002'
    vg_grade_id = '20000000-0000-0000-0000-000000000003'
    vg_thickness_id = '20000000-0000-0000-0000-000000000004'

    variation_groups = {
        'ขนาด': vg_size_id,
        'สี': vg_color_id,
        'เกรด': vg_grade_id,
        'ความหนา': vg_thickness_id,
    }

    # Collect all unique variation entries across all products
    all_sizes = set()
    all_colors = set()
    all_grades = set()
    all_thicknesses = set()

    for p in all_products.values():
        all_sizes.update(p['sizes'])
        all_colors.update(p['colors'])
        all_grades.update(p['grades'])
        all_thicknesses.update(p['thicknesses'])

    # Generate UUIDs for each variation entry
    entry_ids = {}  # (group_name, value) -> uuid

    def get_entry_id(group_name, value):
        key = (group_name, value)
        if key not in entry_ids:
            entry_ids[key] = str(uuid.uuid4())
        return entry_ids[key]

    # Pre-generate all entry IDs
    for v in sorted(all_sizes): get_entry_id('ขนาด', v)
    for v in sorted(all_colors): get_entry_id('สี', v)
    for v in sorted(all_grades): get_entry_id('เกรด', v)
    for v in sorted(all_thicknesses): get_entry_id('ความหนา', v)

    # Generate SQL
    sql_lines = []
    sql_lines.append('-- Auto-generated from xlsx: Seed ~39 products with variations')
    sql_lines.append('-- Generated by scripts/seed-products-from-xlsx.py')
    sql_lines.append('')

    # 1. Variation groups
    sql_lines.append('-- ============================================================')
    sql_lines.append('-- VARIATION GROUPS')
    sql_lines.append('-- ============================================================')
    sql_lines.append('')
    sql_lines.append('INSERT INTO variation_groups (id, name) VALUES')
    vg_values = []
    for name, gid in variation_groups.items():
        vg_values.append(f"  ('{gid}', {escape_sql(name)})")
    sql_lines.append(',\n'.join(vg_values))
    sql_lines.append('ON CONFLICT (id) DO NOTHING;')
    sql_lines.append('')

    # 2. Variation entries
    sql_lines.append('-- ============================================================')
    sql_lines.append('-- VARIATION ENTRIES')
    sql_lines.append('-- ============================================================')
    sql_lines.append('')

    group_entries = {
        'ขนาด': sorted(all_sizes),
        'สี': sorted(all_colors),
        'เกรด': sorted(all_grades),
        'ความหนา': sorted(all_thicknesses),
    }

    entry_values = []
    for group_name, values in group_entries.items():
        gid = variation_groups[group_name]
        for i, val in enumerate(values):
            eid = get_entry_id(group_name, val)
            entry_values.append(f"  ('{eid}', '{gid}', {escape_sql(val)}, {i})")

    if entry_values:
        sql_lines.append('INSERT INTO variation_entries (id, group_id, label, sort_order) VALUES')
        sql_lines.append(',\n'.join(entry_values))
        sql_lines.append('ON CONFLICT (id) DO NOTHING;')
        sql_lines.append('')

    # 3. Products
    sql_lines.append('-- ============================================================')
    sql_lines.append('-- PRODUCTS')
    sql_lines.append('-- ============================================================')
    sql_lines.append('')

    product_ids = {}  # key -> uuid
    product_values = []
    sort_order = 0

    for key, p in all_products.items():
        pid = str(uuid.uuid4())
        product_ids[key] = pid

        slug = slugify(p['name_th'])
        # Ensure slug uniqueness by appending category if needed
        code = p['name_en'].upper().replace(' ', '-')[:20] if p['name_en'] else slug.upper()[:20]
        sku = list(p['skus'])[0] if p['skus'] else code

        # Use first description
        desc = list(p['descriptions'])[0] if p['descriptions'] else p['name_th']

        product_values.append(
            f"  ('{pid}', {escape_sql(code)}, {escape_sql(sku)}, {escape_sql(p['name_th'])}, "
            f"{escape_sql(slug)}, '{p['type']}', {escape_sql(p['category'])}, "
            f"{escape_sql(desc)}, true, false, {sort_order})"
        )
        sort_order += 1

    sql_lines.append('INSERT INTO products (id, code, sku, name, slug, type, category, description, published, recommended, sort_order) VALUES')
    sql_lines.append(',\n'.join(product_values))
    sql_lines.append('ON CONFLICT (id) DO NOTHING;')
    sql_lines.append('')

    # 4. Product variation links
    sql_lines.append('-- ============================================================')
    sql_lines.append('-- PRODUCT VARIATION LINKS')
    sql_lines.append('-- ============================================================')
    sql_lines.append('')

    link_values = []
    for key, p in all_products.items():
        pid = product_ids[key]

        # Only link variations when there are 2+ values (single value = not a real variation)
        if len(p['sizes']) >= 2:
            for size in p['sizes']:
                eid = get_entry_id('ขนาด', size)
                link_values.append(f"  ('{str(uuid.uuid4())}', '{pid}', '{vg_size_id}', '{eid}')")

        if len(p['colors']) >= 2:
            for color in p['colors']:
                eid = get_entry_id('สี', color)
                link_values.append(f"  ('{str(uuid.uuid4())}', '{pid}', '{vg_color_id}', '{eid}')")

        if len(p['grades']) >= 2:
            for grade in p['grades']:
                eid = get_entry_id('เกรด', grade)
                link_values.append(f"  ('{str(uuid.uuid4())}', '{pid}', '{vg_grade_id}', '{eid}')")

        if len(p['thicknesses']) >= 2:
            for thickness in p['thicknesses']:
                eid = get_entry_id('ความหนา', thickness)
                link_values.append(f"  ('{str(uuid.uuid4())}', '{pid}', '{vg_thickness_id}', '{eid}')")

    if link_values:
        sql_lines.append('INSERT INTO product_variation_links (id, product_id, group_id, entry_id) VALUES')
        sql_lines.append(',\n'.join(link_values))
        sql_lines.append('ON CONFLICT (product_id, entry_id) DO NOTHING;')
        sql_lines.append('')

    # Summary
    sql_lines.append(f'-- Summary: {len(all_products)} products, {len(entry_ids)} variation entries, {len(link_values)} links')

    sql = '\n'.join(sql_lines)

    with open(OUTPUT_PATH, 'w', encoding='utf-8') as f:
        f.write(sql)

    print(f'Generated {OUTPUT_PATH}')
    print(f'  Products: {len(all_products)}')
    print(f'  Variation groups: {len(variation_groups)}')
    print(f'  Variation entries: {len(entry_ids)}')
    print(f'  Product-variation links: {len(link_values)}')

    # Print product summary
    print('\nProducts:')
    for key, p in all_products.items():
        variants = []
        if p['sizes']: variants.append(f"{len(p['sizes'])} sizes")
        if p['colors']: variants.append(f"{len(p['colors'])} colors")
        if p['grades']: variants.append(f"{len(p['grades'])} grades")
        if p['thicknesses']: variants.append(f"{len(p['thicknesses'])} thicknesses")
        vstr = ', '.join(variants) if variants else 'no variants'
        print(f"  [{p['type'][:5]}] {p['category']} > {p['name_th']} ({vstr})")


if __name__ == '__main__':
    main()
