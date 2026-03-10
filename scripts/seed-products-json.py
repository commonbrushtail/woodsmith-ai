"""Generate JSON files for seeding products from xlsx."""
import openpyxl
import uuid
import json
import re

XLSX_PATH = 'Vanachai website map 040326/Vanachai website map/Website map - photo.xlsx'

CATEGORY_TYPE = {
    'โต๊ะ': 'decoration', 'ประตู': 'decoration', 'ไม้โพลีเอสเตอร์': 'construction',
    'MDF': 'construction', 'PB': 'construction', 'OSB': 'construction',
    'ไม้อัด': 'construction', 'ไม้บันได': 'decoration', 'ไม้แบบ': 'construction',
    'ไม้พื้น': 'decoration', 'ไม้ฝาตกแต่ง': 'decoration',
}

SLUG_MAP = {
    'โต๊ะ': 'table', 'ประตูเมลามีน': 'melamine-door',
    'ประตูเมลามีนพร้อมวงกบ': 'melamine-door-with-frame',
    'ประตูเมลามีนกันน้ำ': 'ultra-melamine-door',
    'ประตูลูกฟัก': 'hdf-door', 'ประตูลูกฟักพร้อมวงกบ': 'hdf-door-with-frame',
    'ประตูHMR': 'hmr-door', 'ประตูปิดPVC': 'pvc-door',
    'ประตูไม้อัด': 'plywood-door', 'ประตูสักประดิษฐ์': 'recompose-teak-door',
    'หน้าประตู': 'doorskin', 'ไม้โพลีเอสเตอร์': 'polyester-board',
    'ไม้ PB': 'pb-board', 'ไม้ PB ปิดผิวเมลามีน': 'pb-melamine',
    'ไม้ OSB2 ภายใน': 'osb2-interior', 'ไม้ OSB2 ทนชื้น': 'osb2-moisture',
    'ไม้ OSB2 ภายนอก': 'osb2-exterior', 'ไม้ OSB3 ภายใน': 'osb3-interior',
    'ไม้ OSB3 ภายนอก': 'osb3-exterior', 'ไม้พื้นลามิเนต': 'laminate-flooring',
    'ไม้พื้นลามิเนตปลอดสารฟอมัลดีไฮด์': 'formaldehyde-free-laminate',
    'ไม้พื้น Veneer Pattern': 'veneer-pattern-flooring',
    'ไม้พื้น Veneer': 'veneer-flooring',
    'ไม้พื้น Hybrid HMR': 'hybrid-hmr-flooring',
    'ไม้พื้น Hybrid Ultra': 'hybrid-ultra-flooring',
    'ไม้ MDF': 'mdf-board', 'ไม้ MDF ปิดผิวเมลามีน': 'mdf-melamine',
    'ไม้ MDF HMR V313': 'mdf-hmr-v313', 'ไม้ MDF HMR V70': 'mdf-hmr-v70',
    'ไม้ฝาตกแต่ง US': 'decorative-panel-us', 'ไม้ฝาตกแต่ง AR': 'decorative-panel-ar',
    'ไม้ฝา HDF ปิดPVC': 'hdf-pvc-panel', 'ไม้ฝา LVL ลายSaw Cut': 'lvl-saw-cut-panel',
    'ไม้ฝา LVL ลายสน': 'lvl-pine-panel', 'ไม้อัด': 'plywood-board',
    'ไม้อัด ภายใน': 'plywood-interior', 'ไม้อัด ภายนอก': 'plywood-exterior',
    'ไม้แบบ': 'shuttering', 'ไม้บันได': 'stair-wood',
}

def slugify(text):
    if text in SLUG_MAP:
        return SLUG_MAP[text]
    s = text.lower().strip()
    s = re.sub(r'[^a-z0-9\s-]', '', s)
    s = re.sub(r'[\s]+', '-', s)
    return s.strip('-') or f'product-{uuid.uuid4().hex[:8]}'

def main():
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    sheets = ['โต๊ะ', 'ประตู', 'ไม้โพลีเอสเตอร์', 'MDF', 'PB', 'OSB',
              'ไม้อัด', 'ไม้บันได', 'ไม้แบบ', 'ไม้พื้น', 'ไม้ฝาตกแต่ง']

    all_products = {}
    for sheet_name in sheets:
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            name_th = row[0]
            if not name_th:
                continue
            name_th = str(name_th).strip()
            if name_th.startswith('*'):
                continue
            key = (sheet_name, name_th)
            if key not in all_products:
                all_products[key] = {
                    'name_th': name_th,
                    'name_en': str(row[1]).strip() if row[1] else '',
                    'category': sheet_name,
                    'type': CATEGORY_TYPE.get(sheet_name, 'construction'),
                    'sizes': set(), 'colors': set(), 'grades': set(), 'thicknesses': set(),
                    'descriptions': set(), 'skus': set(),
                }
            p = all_products[key]
            if row[2]: p['sizes'].add(str(row[2]).strip())
            if row[7]: p['colors'].add(str(row[7]).strip())
            if row[4]: p['grades'].add(str(row[4]).strip())
            if row[8]: p['thicknesses'].add(str(row[8]).strip())
            if row[3]: p['descriptions'].add(str(row[3]).strip())
            if row[6]: p['skus'].add(str(row[6]).strip())

    # Variation groups
    vgroups = {
        'ขนาด': '20000000-0000-0000-0000-000000000001',
        'สี': '20000000-0000-0000-0000-000000000002',
        'เกรด': '20000000-0000-0000-0000-000000000003',
        'ความหนา': '20000000-0000-0000-0000-000000000004',
    }

    # Collect unique entries per group
    entry_ids = {}
    def get_eid(group, val):
        k = (group, val)
        if k not in entry_ids:
            entry_ids[k] = str(uuid.uuid4())
        return entry_ids[k]

    # Build variation entries
    all_entries = []
    seen = set()
    for p in all_products.values():
        for s in p['sizes']: get_eid('ขนาด', s)
        for c in p['colors']: get_eid('สี', c)
        for g in p['grades']: get_eid('เกรด', g)
        for t in p['thicknesses']: get_eid('ความหนา', t)

    sort_counters = {'ขนาด': 0, 'สี': 0, 'เกรด': 0, 'ความหนา': 0}
    for (group, val), eid in sorted(entry_ids.items()):
        all_entries.append({
            'id': eid,
            'group_id': vgroups[group],
            'label': val,
            'sort_order': sort_counters[group],
        })
        sort_counters[group] += 1

    # Build products and links
    products_list = []
    links_list = []
    sort_order = 0

    for key, p in all_products.items():
        pid = str(uuid.uuid4())
        slug = slugify(p['name_th'])
        code = p['name_en'].upper().replace(' ', '-')[:20] if p['name_en'] else slug.upper()[:20]
        sku = list(p['skus'])[0] if p['skus'] else code
        desc = list(p['descriptions'])[0] if p['descriptions'] else p['name_th']

        products_list.append({
            'id': pid, 'code': code, 'sku': sku, 'name': p['name_th'],
            'slug': slug, 'type': p['type'], 'category': p['category'],
            'description': desc, 'published': True, 'recommended': False,
            'sort_order': sort_order,
        })
        sort_order += 1

        # Only link variations with 2+ values
        if len(p['sizes']) >= 2:
            for v in p['sizes']:
                links_list.append({'id': str(uuid.uuid4()), 'product_id': pid, 'group_id': vgroups['ขนาด'], 'entry_id': get_eid('ขนาด', v)})
        if len(p['colors']) >= 2:
            for v in p['colors']:
                links_list.append({'id': str(uuid.uuid4()), 'product_id': pid, 'group_id': vgroups['สี'], 'entry_id': get_eid('สี', v)})
        if len(p['grades']) >= 2:
            for v in p['grades']:
                links_list.append({'id': str(uuid.uuid4()), 'product_id': pid, 'group_id': vgroups['เกรด'], 'entry_id': get_eid('เกรด', v)})
        if len(p['thicknesses']) >= 2:
            for v in p['thicknesses']:
                links_list.append({'id': str(uuid.uuid4()), 'product_id': pid, 'group_id': vgroups['ความหนา'], 'entry_id': get_eid('ความหนา', v)})

    # Write JSON
    data = {
        'variation_groups': [{'id': v, 'name': k} for k, v in vgroups.items()],
        'variation_entries': all_entries,
        'products': products_list,
        'product_variation_links': links_list,
    }

    with open('scripts/seed-data.json', 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    print(f'Generated scripts/seed-data.json')
    print(f'  Products: {len(products_list)}')
    print(f'  Variation entries: {len(all_entries)}')
    print(f'  Links: {len(links_list)}')

if __name__ == '__main__':
    main()
